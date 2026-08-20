import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path
 
 
def four_pl(logx, A, B, logEC50, D):
    return D + (A - D) / (1 + 10 ** ((logEC50 - logx) * B))
 
 
def simulate_elisa_data(seed=42):
    rng = np.random.default_rng(seed)
 
    # 8-point standard curve: 32 ng/mL, 2-fold serial dilution
    STD = [32 * (0.5 ** i) for i in range(8)]
    A_true, B_true, C_true, D_true = 3.2, 1.2, 3.8, 0.03
 
    def noisy_duplicate(x_vals, A, B, C, D, noise=0.03):
        base = four_pl(np.log10(np.array(x_vals)), A, B, np.log10(C), D)
        rep1 = base * (1 + rng.normal(0, noise, len(base)))
        rep2 = base * (1 + rng.normal(0, noise, len(base)))
        return rep1, rep2
 
    std_r1, std_r2 = noisy_duplicate(STD, A_true, B_true, C_true, D_true)
 
    sample_names = ["A1", "A2", "A3", "A4"]
    timepoints = [
        "Assay", "Assay 5x", "Assay 10x", "Assay 100x",
        "Assay 500x", "Assay 1000x", "TestSample 10x", "TestSample 100x",
    ]
 
    sample_cols = {}
    for name in sample_names:
        base_curve = rng.uniform(0.05, 3.5, size=8)
        rep1 = base_curve * (1 + rng.normal(0, 0.05, 8))
        rep2 = base_curve * (1 + rng.normal(0, 0.05, 8))
        sample_cols[name] = (rep1, rep2)
 
    blank_col_a = rng.uniform(0.02, 0.03, 8)
    blank_col_b = rng.uniform(0.02, 0.03, 8)

    #Control sample
    CONTROL_OD = 1.6
    blank_col_a[7] = CONTROL_OD * (1 + rng.normal(0,0.03))
    blank_col_b[7] = CONTROL_OD * (1 + rng.normal(0,0.03))

 
    # 12 data columns total: STD (2) + A1..A4 (8) + blank (2)
    od_columns = [std_r1, std_r2]
    for name in sample_names:
        od_columns.extend(sample_cols[name])
    od_columns.extend([blank_col_a, blank_col_b * (1 + rng.normal(0, 0.05, 8))])
 
    label_columns = [f"STD {STD[i]:.4f}".replace(".", ",") for i in range(8)], None
    return STD, od_columns, sample_names, timepoints
 
 
def build_workbook(output_path, wanted_od="OD(450) - OD(570)"):
    STD, od_columns, sample_names, timepoints = simulate_elisa_data()
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Results by plate"
 
    # --- Junk header rows above the OD block (mimics real instrument export) ---
    ws["A1"] = "Instrument: Synthetic Plate Reader v1.0"
    ws["A2"] = "Export generated for testing purposes only"
 
    od_label_row = 40  # arbitrary offset, matches the "row 41" position in your screenshot
 
    # Row: OD channel label
    ws.cell(row=od_label_row, column=1, value=wanted_od)
 
    # Row: plate column numbers (1..12) -- this row + first col get trimmed by load_workbook_grid
    header_row = od_label_row + 1
    for col_idx in range(1, 13):
        ws.cell(row=header_row, column=col_idx + 2, value=col_idx)
 
    # 8 OD data rows, columns C..N (12 data columns), with plate row/letter labels in A/B
    data_start_row = header_row + 1
    row_letters = list("ABCDEFGH")
    for r in range(8):
        excel_row = data_start_row + r
        ws.cell(row=excel_row, column=1, value=1)  # plate "row group" index
        ws.cell(row=excel_row, column=2, value=row_letters[r])
        for c in range(12):
            ws.cell(row=excel_row, column=c + 3, value=round(float(od_columns[c][r]), 3))
 
    # 8 label rows directly below
    label_start_row = data_start_row + 8
    for r in range(8):
        excel_row = label_start_row + r
        ws.cell(row=excel_row, column=1, value=None)
        ws.cell(row=excel_row, column=2, value=None)
 
        # STD columns (2 duplicate columns, only first carries the label)
        std_label = f"STD {STD[r]:.4f}".replace(".", ",")
        ws.cell(row=excel_row, column=3, value=std_label)
        ws.cell(row=excel_row, column=4, value=None)
 
        # Sample columns
        col = 5
        for name in sample_names:
            ws.cell(row=excel_row, column=col, value=f"{name} {timepoints[r]}")
            ws.cell(row=excel_row, column=col + 1, value=None)
            col += 2
 
        # Blank columns (last row also gets a Control label, like your real data)
        if r < 7:
            ws.cell(row=excel_row, column=col, value="blank")
        else:
            ws.cell(row=excel_row, column=col, value="Control")
        ws.cell(row=excel_row, column=col + 1, value=None)
 
    # --- Color-scale heatmap on the OD data block, like the real instrument export ---
    data_range = f"C{data_start_row}:N{data_start_row + 7}"
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",   # red = low
        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow = mid
        end_type="max", end_color="63BE7B",       # green = high
    )
    ws.conditional_formatting.add(data_range, rule)
 
    # Light styling for readability
    ws.cell(row=od_label_row, column=1).font = Font(bold=True)
    for col_idx in range(1, 13):
        cell = ws.cell(row=header_row, column=col_idx + 2)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
 
    wb.save(output_path)
    print(f"Saved synthetic plate export to: {output_path}")
    print(f"OD channel row: {od_label_row} | Sheet name: 'Results by plate'")
 
 
if __name__ == "__main__":
    out = Path("synthetic_elisa_plate.xlsx")
    build_workbook(out)